import os
from openpyxl import Workbook, load_workbook

COLUMNS = ["Name", "Email", "Skill Level", "Score", "Result", "DateTime", "Unique ID"]

def ensure_results_file(path: str):
    os.makedirs(os.path.dirname(path), exist_ok=True)
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Results"
        ws.append(COLUMNS)
        wb.save(path)

def append_result(path: str, name: str, email: str, level: int, score: int, result: str, timestamp: str, unique_id: str):
    if not os.path.exists(path):
        ensure_results_file(path)
    wb = load_workbook(path)
    ws = wb.active
    ws.append([name, email, level, score, result, timestamp, unique_id])
    wb.save(path)
